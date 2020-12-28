import pickle
from num2words import num2words
import openpyxl
from datetime import *
from data import config

import math
from PIL import Image


class Invoice():
    """Класс Invoice используется для создания Счетов на оплату

        Основное применение - формирование Счета и экспорт данных в формат xlsx

        Note:
            Все атрибуты заданы по-умолчанию

        Attributes
        ----------
        quantity : int
            количество товара или услуги
        name_of_services : str
            название услуги
        customer : str
            наименование покупателя (организации)
        executor: str
            Название и реквизиты исполнителя
        units: str
            еденицы измерения (шт, м.кв.)
        price: int
            цена одной еденицы услуги
        nds: str
            прописывается поле с НДС
        summa: int
            считается автоматически
        names: int
            общее кол-во наименований
        director: str
            директор
        bookseeper: str
            бухгалтер

        Methods
        -------
        exel_generator()
            Создает эксель файл со Счетом
        summa_propisyu ()
            Приводит число к принятому в документах текстовому отображению
        names_string_result()
            Выводит строку 'Счет на оплату №...от....'
        date_words()
            Формирует дату в стандартном для счета представлении
        number()
            Формирует номер счета, прибавляя еденицу к данным из счетчика
        """

    def __str__(self):  # Магический дандр метод форматирования вывода print(p): Дело(Название=1, Сделано?=1)
        return f'Документ: {self.number()} {self.date_words()}\n' \
               f'   Заказчик: {self.customer}\n' \
               f'   Исполнитель: {self.executor}\n' \
               f'   Сумма: {self.summa} рублей \n' \
               f'   Основание: {self.name_of_services}\n'

    def exel_generator(self):

        from openpyxl import load_workbook
        myimage = Image.open(f'{config.PICTURE_PATH}stamp_neoluxe.png')
        myimage.load()
        myimage2 = Image.open(f'{config.PICTURE_PATH}autograph.png')
        myimage2.load()
        myimage3 = Image.open(f'{config.PICTURE_PATH}autograph2.png')
        myimage3.load()
        myimage4 = Image.open(f'{config.PICTURE_PATH}logo.png')
        myimage4.load()
        wb = load_workbook(config.EXEL_PATH + 'example2.xlsx')
        wb_filename = f'{config.EXEL_PATH}invoice_{self.summa}_{self.number()}.xlsx'
        ws = wb["Заготовка счета"]
        ws['U22'] = self.quantity  # Количество
        ws['D22'] = self.name_of_services  # Товар
        ws['AB22'] = self.price  # Цена
        ws['Y22'] = self.units  # Еденицы измерения
        ws['F19'] = self.customer  # Покупатель
        ws['AH22'] = self.summa  # Сумма
        ws['AH24'] = self.summa  # Итого
        ws['AH26'] = self.summa  # Всего к оплате
        ws['B27'] = self.names_string_result()  # Всего наименований
        ws['B28'] = self.summa_propisyu()  # Сумма прописью
        ws['B13'] = f'{self.number()}{self.date_words()}'  # Номер счета
        img1 = openpyxl.drawing.image.Image(config.STAMP)
        img1.anchor = ws.cell(row=29, column=9).coordinate
        ws.add_image(img1)
        img2 = openpyxl.drawing.image.Image(config.AUTOGRAPH)
        img2.anchor = ws.cell(row=30, column=7).coordinate
        ws.add_image(img2)
        img3 = openpyxl.drawing.image.Image(config.AUTOGRAPH2)
        img3.anchor = ws.cell(row=30, column=28).coordinate
        ws.add_image(img3)
        img4 = openpyxl.drawing.image.Image(config.LOGO)
        img4.anchor = ws.cell(row=1, column=2).coordinate
        ws.add_image(img4)
        wb.save(filename=wb_filename)

    def summa_propisyu(self):
        kops = int(self.price) * float(self.quantity)
        try:
            kopeyki = math.modf(kops)[0]
        except:
            breakpoint()
        rubli = math.modf(float(self.summa))[1]
        rubliround = int(rubli)
        word3 = num2words(str(rubliround), lang='ru')
        wrd31 = (word3.capitalize())
        newkopeyki = str(kopeyki)
        newkopeyki2 = newkopeyki.replace('.', "")
        return f'{str(wrd31)} рублей {newkopeyki2} копеек. НДС 0%."'

    def names_string_result(self):
        return f'Всего наименований {self.names}, на сумму {self.summa} руб.'

    def number(self):
        with open(config.SCHETCHIK_PATH, 'rb') as f:
            loaded_data = pickle.load(f)  # Загружаем экземпляр класса
        return f'Счет на оплату № {int(loaded_data["value"]) + 1} от '

    def date_words(self):
        a = ['Января', 'Февраля', 'Марта', 'Апреля', 'Мая', 'Июня', 'Июля', 'Августа', 'Сентября', 'Октября', 'Ноября',
             'Декабря']
        a1 = ['01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14', '15', '16', '17',
              '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31']
        d = datetime.today()
        g = d.day
        c = d.month
        h = d.year

        return (f'{a1[g - 1]} {a[c - 1]} {str(h)}')

    def __init__(self,
                 customer='ИП Телелинский Дмитрий Владимирович  , ИНН  - , КПП   -, 350063, город Краснодар, ул.Кубанская Набережная, д.37, кв.140',
                 executor='ООО "Неолюкс", ИНН 6658215373, КПП 231001001, 350010, г. Краснодар, ул. Зиповская, д. 5, Телефон: +7 (978) 074-08-54 ',
                 name_of_services='Дизайн проект по Договору', units='кв.м.', quantity='100', price='400',
                 nds='Без НДС', names=1,
                 director='Кандалов В.А.', bookkeeper='Кандалов В.А.'):
        # self.number = number  # Номер счета (достается из счетчика)
        self.customer = customer  # Заказчик (достается из словаря с заказчиками или из БД)
        self.executor = executor  # Исполнитель
        self.name_of_services = name_of_services  # Наименование услуг
        self.units = units  # Еденицы измерения
        self.quantity = quantity  # Количество
        self.price = price  # Цена
        self.summa = '{0:.2f}'.format(float(self.price) * float(self.quantity))  # Сумма
        self.data = datetime.now()  # Дата счета
        self.nds = nds
        self.names = names
        self.director = director
        self.bookkeeper = bookkeeper


def Schetchic_Saver(schetchik):
    with open(config.SCHETCHIK_PATH, "wb") as f:
        pickle.dump(schetchik, f, protocol=pickle.HIGHEST_PROTOCOL)


def Schetchic_clear():
    with open(config.SCHETCHIK_PATH, 'rb') as f:
        loaded_data = pickle.load(f)  # Загружаем экземпляр класса
    loaded_data['value'] = 0
    Schetchic_Saver(loaded_data)


def Schetchic_Pusher():
    with open(config.SCHETCHIK_PATH, 'rb') as f:
        loaded_data = pickle.load(f)  # Загружаем экземпляр класса
    loaded_data['value'] += 1
    Schetchic_Saver(loaded_data)


def InvoceMaker():
    newInvoice = Invoice(price=500, quantity=989)
    newInvoice.exel_generator()
    Schetchic_Pusher()
    newInvoice2 = Invoice(price=600, quantity=700)
    newInvoice2.exel_generator()


newInvoice = Invoice(price=50)
print(newInvoice)

# Schetchic_clear()
